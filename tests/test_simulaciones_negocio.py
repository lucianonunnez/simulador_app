"""
Regresión contra simulaciones REALES del negocio (reconciliación 0.0000%).

Cada archivo es un workbook de negociación ya resuelto por el equipo (hoja
'Simulación' con Cantidad, Valor actual/propuesto por prestación, y la cabecera
'Solicitado/Propuesto' con el Impacto año/mes/%). Este test alimenta la lógica
PURA de la app (core.simulator) con esos insumos y verifica que reproduce los
mismos números, por DOS caminos independientes:

  Path 1 — impact_metrics() reproduce el Impacto (año/mes/%) de la cabecera.
  Path 2 — apply_simulation() reproduce, fila a fila, el Valor Ofrecido y el
           Consumo Simulado del negocio a partir de la CONFIG de aumento
           (plano, o 'capas' con override por prestación).

Validado en la sesión de testing con 3 simulaciones (HI jun-26, 1130 feb-26,
1127 dic-25) con desvío máximo ~1e-15 (ruido de punto flotante).

DATOS SENSIBLES: los .xlsx NO se versionan (precios de prestadores). El test se
SALTEA si no están. Para correrlo, dejá los workbooks en
`tests/fixtures/simulaciones/` (o exportá SIM_FIXTURES_DIR apuntando a la
carpeta) con estos nombres (o ajustá CASOS):

    Simulacion_Hospital_Italiano_jun26.xlsx
    Simulacion_1130_Feb26.xlsx
    Simulacion_1127_Dic25.xlsx
"""

from __future__ import annotations

import os
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from core.simulator import apply_simulation, impact_metrics

# ----------------------------------------------------------------------------
# Casos: archivo + headline esperado (escenario Propuesto) + config de aumento
# ----------------------------------------------------------------------------
CASOS = {
    "HI_jun26": {
        "archivo": "Simulacion_Hospital_Italiano_jun26.xlsx",
        "pauta": 2.1,
        "impacto_pct": 0.020349819019435555,
        "impacto_anio": 316594963.51207733,
        "impacto_mes": 26382913.626006443,
        # Diferenciales de laboratorio a 0%, resto 2.1% -> modo 'capas' con
        # override por prestación (descubierto del propio workbook).
        "mode": "capas",
        "flat": 2.1,
    },
    "Sim1130_Feb26": {
        "archivo": "Simulacion_1130_Feb26.xlsx",
        "pauta": 2.2,
        "impacto_pct": 0.021999999999999353,
        "impacto_anio": 325840125.86333084,
        "impacto_mes": 27153343.821944237,
        "mode": "plano",
        "flat": 2.2,
    },
    "Sim1127_Dic25": {
        "archivo": "Simulacion_1127_Dic25.xlsx",
        "pauta": 1.8,
        "impacto_pct": 0.018999999999994577,
        "impacto_anio": 271784065.809885,
        "impacto_mes": 22648672.150823753,
        "mode": "plano",
        "flat": 1.9,
    },
}


def _fixtures_dir() -> Path:
    env = os.environ.get("SIM_FIXTURES_DIR")
    if env:
        return Path(env)
    return Path(__file__).parent / "fixtures" / "simulaciones"


def _ruta(caso: dict) -> Path | None:
    p = _fixtures_dir() / caso["archivo"]
    return p if p.exists() else None


# ----------------------------------------------------------------------------
# Extracción de la hoja 'Simulación' (tolera las 3 variantes de encabezado)
# ----------------------------------------------------------------------------
def _extraer_simulacion(path: Path) -> pd.DataFrame:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hoja = next(s for s in wb.sheetnames if "imula" in s.lower())
    filas = list(wb[hoja].iter_rows(values_only=True))
    wb.close()

    def _norm(r):
        return [str(c).strip() if c is not None else "" for c in r]

    h = next(i for i, r in enumerate(filas) if "Cantidad CM" in _norm(r))
    header = _norm(filas[h])

    def idxs(*nombres):
        out = []
        for j, col in enumerate(header):
            if col.lower() in {n.lower() for n in nombres}:
                out.append(j)
        return out

    def i1(*n):
        xs = idxs(*n)
        return xs[0] if xs else None

    # 'Valor actual/solicitado/propuesto' aparecen DOS veces: 1ra = unitario,
    # 2da = Q×P (el producto que ya calculó el negocio).
    va = idxs("Valor actual", "Valor Actual")
    vp = idxs("Valor propuesto")
    ci, nomc = i1("Cantidad CM"), i1("Nomenclador")
    tipc, pidc = i1("Tipo Clase CM"), i1("idPrestacion", "Cod")
    pauta = i1("Pauta/No pauta", "Pauta / No pauta")

    def num(r, j):
        if j is None:
            return np.nan
        try:
            return float(r[j])
        except (TypeError, ValueError):
            return np.nan

    recs = []
    for r in filas[h + 1:]:
        if r is None or num(r, ci) != num(r, ci):  # Cantidad NaN -> saltar
            continue
        recs.append({
            "tipo": r[tipc] if tipc is not None else None,
            "nomen": r[nomc] if nomc is not None else None,
            "pid": num(r, pidc),
            "pauta": str(r[pauta]).strip() if pauta is not None and r[pauta] is not None else None,
            "cant": num(r, ci),
            "vact": num(r, va[0]) if len(va) > 0 else np.nan,
            "vprop": num(r, vp[0]) if len(vp) > 0 else np.nan,
            "qxp_prop": num(r, vp[1]) if len(vp) > 1 else np.nan,
        })
    return pd.DataFrame(recs)


def _universo_pauta(df: pd.DataFrame) -> pd.DataFrame:
    """Universo simulable: filas 'Pauta' con valor actual y propuesto válidos."""
    d = df[df["pauta"] == "Pauta"].dropna(subset=["vact", "vprop", "qxp_prop"]).copy()
    return d[d["vact"] > 0].reset_index(drop=True)


# ----------------------------------------------------------------------------
# Tests
# ----------------------------------------------------------------------------
@pytest.mark.parametrize("nombre", list(CASOS))
def test_impact_metrics_reproduce_headline_del_negocio(nombre):
    """Path 1: impact_metrics() == Impacto (año/mes/%) de la cabecera del workbook."""
    caso = CASOS[nombre]
    ruta = _ruta(caso)
    if ruta is None:
        pytest.skip(f"fixture ausente: {caso['archivo']} (datos sensibles, no versionados)")

    d = _universo_pauta(_extraer_simulacion(ruta))
    sim = pd.DataFrame({
        "Consumo Ideal": d["cant"] * d["vact"],
        "Consumo Simulado": d["cant"] * d["vprop"],
    })
    m = impact_metrics(sim, pauta_pct=caso["pauta"], n_meses=12)

    assert m["impacto_pct"] == pytest.approx(caso["impacto_pct"], abs=1e-9)
    assert m["impacto"] == pytest.approx(caso["impacto_anio"], rel=1e-9)
    assert m["impacto_mensual"] == pytest.approx(caso["impacto_mes"], rel=1e-9)


@pytest.mark.parametrize("nombre", list(CASOS))
def test_apply_simulation_reproduce_valor_ofrecido_fila_a_fila(nombre):
    """Path 2: apply_simulation() reconstruye, desde la CONFIG de aumento, el
    Valor Ofrecido y el Consumo Simulado que el negocio cargó fila a fila."""
    caso = CASOS[nombre]
    ruta = _ruta(caso)
    if ruta is None:
        pytest.skip(f"fixture ausente: {caso['archivo']} (datos sensibles, no versionados)")

    d = _universo_pauta(_extraer_simulacion(ruta))
    merged = pd.DataFrame({
        "Prestacion ID": pd.to_numeric(d["pid"], errors="coerce").astype("Int64"),
        "Nomenclador": d["nomen"],
        "Tipo Clase CM": d["tipo"],
        "Cantidad CM": d["cant"],
        "Valor Convenido a HOY": d["vact"],
    })

    pres_pcts = None
    if caso["mode"] == "capas":
        # Override por prestación descubierto del propio workbook: las que NO
        # están al % plano (los diferenciales a 0%).
        eff = ((d["vprop"] / d["vact"] - 1) * 100).round(4)
        dev = d[eff.round(2) != round(caso["flat"], 2)]
        pres_pcts = {
            int(p): float(e)
            for p, e in zip(pd.to_numeric(dev["pid"], errors="coerce"), eff[dev.index])
        }

    out = apply_simulation(
        merged, months=1, mode=caso["mode"], flat_pct=caso["flat"],
        prestacion_pcts=pres_pcts,
    )

    rel_vo = np.abs(out["Valor Ofrecido"].to_numpy() - d["vprop"].to_numpy()) / np.where(
        np.abs(d["vprop"].to_numpy()) > 0, np.abs(d["vprop"].to_numpy()), 1)
    rel_cs = np.abs(out["Consumo Simulado"].to_numpy() - d["qxp_prop"].to_numpy()) / np.where(
        np.abs(d["qxp_prop"].to_numpy()) > 0, np.abs(d["qxp_prop"].to_numpy()), 1)

    assert rel_vo.max() < 1e-6, f"Valor Ofrecido difiere en {(rel_vo > 1e-6).sum()} filas"
    assert rel_cs.max() < 1e-6, f"Consumo Simulado difiere en {(rel_cs > 1e-6).sum()} filas"

    # y el agregado reconcilia con el headline
    m = impact_metrics(out, n_meses=12)
    assert m["impacto"] == pytest.approx(caso["impacto_anio"], rel=1e-6)


def test_capas_override_por_prestacion_pisa_el_plano_general():
    """Regresión SIN fixtures (siempre corre): el patrón de HI jun-26 — plano
    general + un puñado de prestaciones a 0% — se reproduce con 'capas'.
    Blinda el camino de negociación por capas aunque no estén los .xlsx."""
    merged = pd.DataFrame({
        "Prestacion ID": [101, 102, 200, 201],
        "Nomenclador": ["Consultas", "Consultas", "Laboratorio", "Laboratorio"],
        "Tipo Clase CM": ["Ambulatorio"] * 4,
        "Cantidad CM": [10, 5, 8, 4],
        "Valor Convenido a HOY": [1000.0, 2000.0, 500.0, 700.0],
    })
    # General 2.1%, pero dos "diferenciales de laboratorio" quedan a 0%.
    out = apply_simulation(
        merged, months=1, mode="capas", flat_pct=2.1,
        prestacion_pcts={200: 0.0, 201: 0.0},
    ).set_index("Prestacion ID")
    assert out.loc[101, "Valor Ofrecido"] == pytest.approx(1021.0)   # +2.1%
    assert out.loc[102, "Valor Ofrecido"] == pytest.approx(2042.0)   # +2.1%
    assert out.loc[200, "Valor Ofrecido"] == pytest.approx(500.0)    # 0%
    assert out.loc[201, "Valor Ofrecido"] == pytest.approx(700.0)    # 0%
    # Impacto global por debajo del 2.1% por los diferenciales a 0%.
    m = impact_metrics(out, pauta_pct=2.1, n_meses=12)
    assert 0 < m["impacto_pct"] < 0.021
