"""
Validación contra los workbooks de negociación del negocio.

Lógica PURA (sin Streamlit): toma los bytes de un workbook ya resuelto por el
equipo (hoja 'Simulación' con Cantidad, Valor actual/solicitado/propuesto por
prestación y sus Q×P) y recalcula el Impacto con el MISMO motor de la app
(`core.simulator.impact_metrics`). Sirve para mostrar, en la UI, que la app
reproduce el workbook — y por dos rutas internas independientes:

  - "reconstruido": Impacto a partir de Cantidad × Valor (lo que hace la app).
  - "Q×P del negocio": Impacto a partir de las columnas Q×P que el workbook
    ya trae calculadas.

Si ambas rutas coinciden, el workbook es internamente consistente y el motor de
la app lo reproduce. Validado contra 3 simulaciones reales con desvío ~1e-15.

Tolera las variantes de encabezado vistas en los workbooks reales
('Valor actual'/'Valor Actual', 'idPrestacion'/'Cod', 'Pauta/No pauta'/
'Pauta / No pauta').
"""

from __future__ import annotations

from io import BytesIO

import numpy as np
import pandas as pd

from core.simulator import impact_metrics

# Meses de la ventana de liquidación del negocio (para el impacto mensual).
N_MESES_NEGOCIO = 12


def _norm_fila(row) -> list[str]:
    return [str(c).strip() if c is not None else "" for c in row]


def extraer_simulacion(file_bytes: bytes) -> pd.DataFrame:
    """
    Extrae la hoja 'Simulación' de un workbook a un DataFrame por prestación.

    Columnas devueltas: tipo (ámbito), nomen, pid, pauta, cant, vact, vsol,
    vprop (unitarios) y qxp_act/qxp_sol/qxp_prop (los Q×P del negocio).
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        hojas = [s for s in wb.sheetnames if "imula" in s.lower()]
        if not hojas:
            raise ValueError("El archivo no tiene una hoja de 'Simulación'.")
        filas = list(wb[hojas[0]].iter_rows(values_only=True))
    finally:
        wb.close()

    h = next(
        (i for i, r in enumerate(filas) if "Cantidad CM" in _norm_fila(r)), None
    )
    if h is None:
        raise ValueError("No se encontró la fila de encabezados ('Cantidad CM').")
    header = _norm_fila(filas[h])

    def idxs(*nombres):
        objetivo = {n.lower() for n in nombres}
        return [j for j, col in enumerate(header) if col.lower() in objetivo]

    def i1(*n):
        xs = idxs(*n)
        return xs[0] if xs else None

    # Las triplas 'Valor actual/solicitado/propuesto' aparecen DOS veces:
    # 1ra = unitario, 2da = Q×P (producto que ya calculó el negocio).
    va = idxs("Valor actual", "Valor Actual")
    vs = idxs("Valor solicitado")
    vp = idxs("Valor propuesto")
    ci = i1("Cantidad CM")
    nomc, tipc = i1("Nomenclador"), i1("Tipo Clase CM")
    pidc = i1("idPrestacion", "Cod")
    paut = i1("Pauta/No pauta", "Pauta / No pauta")

    def cell(r, j):
        # Las filas de read_only pueden venir más cortas que el encabezado:
        # tratar la celda ausente como vacía, no como IndexError.
        if j is None or j >= len(r):
            return None
        return r[j]

    def num(r, j):
        try:
            return float(cell(r, j))
        except (TypeError, ValueError):
            return np.nan

    recs = []
    for r in filas[h + 1:]:
        if r is None:
            continue
        cant = num(r, ci)
        if cant != cant:  # NaN -> fila de título/subtotal
            continue
        pauta_val = cell(r, paut)
        recs.append({
            "tipo": cell(r, tipc),
            "nomen": cell(r, nomc),
            "pid": num(r, pidc),
            "pauta": str(pauta_val).strip() if pauta_val is not None else None,
            "cant": cant,
            "vact": num(r, va[0]) if len(va) > 0 else np.nan,
            "vsol": num(r, vs[0]) if len(vs) > 0 else np.nan,
            "vprop": num(r, vp[0]) if len(vp) > 0 else np.nan,
            "qxp_act": num(r, va[1]) if len(va) > 1 else np.nan,
            "qxp_sol": num(r, vs[1]) if len(vs) > 1 else np.nan,
            "qxp_prop": num(r, vp[1]) if len(vp) > 1 else np.nan,
        })
    return pd.DataFrame(recs)


def _impacto_escenario(d: pd.DataFrame, col_unit: str, col_qxp: str) -> dict | None:
    """Impacto de un escenario por las dos rutas (reconstruido / Q×P)."""
    valido = d.dropna(subset=["vact", col_unit])
    valido = valido[valido["vact"] > 0]
    if len(valido) == 0:
        return None

    recon = impact_metrics(pd.DataFrame({
        "Consumo Ideal": valido["cant"] * valido["vact"],
        "Consumo Simulado": valido["cant"] * valido[col_unit],
    }), n_meses=N_MESES_NEGOCIO)

    out = {
        "filas": int(len(valido)),
        "impacto_pct": recon["impacto_pct"],
        "impacto": recon["impacto"],
        "impacto_mensual": recon["impacto_mensual"],
        "desvio_qxp": None,
    }

    # Ruta independiente: los Q×P que el workbook ya trae calculados.
    vq = valido.dropna(subset=["qxp_act", col_qxp])
    if len(vq):
        qxp = impact_metrics(pd.DataFrame({
            "Consumo Ideal": vq["qxp_act"],
            "Consumo Simulado": vq[col_qxp],
        }), n_meses=N_MESES_NEGOCIO)
        out["impacto_pct_qxp"] = qxp["impacto_pct"]
        out["desvio_qxp"] = abs(qxp["impacto_pct"] - recon["impacto_pct"])
    return out


def validar_workbook(file_bytes: bytes) -> dict:
    """
    Recalcula el Impacto (Solicitado y Propuesto) de un workbook con el motor
    de la app, por dos rutas independientes.

    Returns:
        {
          "n_filas": int, "n_pauta": int,
          "solicitado": {impacto_pct, impacto, impacto_mensual, desvio_qxp, ...} | None,
          "propuesto":  {...} | None,
        }
    """
    df = extraer_simulacion(file_bytes)
    if df.empty:
        # Hoja 'Simulación' sin filas de datos: resultado vacío con mensaje de
        # negocio en la UI, no un KeyError (df vacío no trae ni las columnas).
        return {"n_filas": 0, "n_pauta": 0, "solicitado": None, "propuesto": None}

    pauta = df[df["pauta"] == "Pauta"] if "pauta" in df.columns else df
    universo = pauta if len(pauta) else df

    return {
        "n_filas": int(len(df)),
        "n_pauta": int(len(pauta)),
        "solicitado": _impacto_escenario(universo, "vsol", "qxp_sol"),
        "propuesto": _impacto_escenario(universo, "vprop", "qxp_prop"),
    }
